import numpy as np
import openpyxl as op
import math


# 计算区间数的直觉模糊数
def qjs_ifn(j):

    q = j[:, 3]                                # 拆分出第四列的区间数
    q1 = np.array([q[0], q[1], q[2], q[3], q[4], q[5], q[6], q[7], q[8], q[9]])          # 根据坦克数量进行变化
    max_dimension2 = np.max(q1, axis=0)        # 取各列最大值
    for row in range(0, q1.shape[0]):          # 计算直觉模糊数

        q1[row, 0] = round(q1[row, 0] / max_dimension2[1], 9)
        q1[row, 1] = round(1 - q1[row, 1] / max_dimension2[1], 9)
    list1 = q1.tolist()

    return list1


# 计算实数的直觉模糊数
def ss_ifn(jz, num1, num2):

    bt = 0.2                                   # 设置模糊因子
    q2 = jz[:, num1:num2]                      # 拆分列
    list2 = q2.tolist()
    max_dimension1 = np.max(q2, axis=0)        # 求各列最大值
    min_dimension1 = np.min(q2, axis=0)        # 求各列最小值
    for column in range(0, q2.shape[0]):       # 求直觉模糊数
        for row in range(0, q2.shape[1]):
            if num1 == 0:
                if row == 1 | row == 2:
                    list2[column][row] = [round(bt * q2[column, row] / max_dimension1[row], 9),
                                          round(bt * (1 - q2[column, row] / max_dimension1[row]), 9)]
                else:
                    list2[column][row] = [round(bt * min_dimension1[row] / q2[column, row], 9),
                                          round(bt * (1 - min_dimension1[row] / q2[column, row]), 9)]
            else:
                if row == 4:
                    list2[column][row] = [round(bt * q2[column, row] / max_dimension1[row], 9),
                                          round(bt * (1 - q2[column, row] / max_dimension1[row]), 9)]
                else:
                    list2[column][row] = [round(bt * min_dimension1[row] / q2[column, row], 9),
                                          round(bt * (1 - min_dimension1[row] / q2[column, row]), 9)]

    return list2


# 求最大最小值
def find_min_max(jz2, num3):

    q4 = jz2[:, num3]                          # 拆分列
    q5 = np.array([q4[0], q4[1], q4[2], q4[3], q4[4], q4[5], q4[6], q4[7], q4[8], q4[9]])       # 根据坦克数量进行变化
    max_uv = np.max(q5, axis=0)                # 求各列最大值
    min_uv = np.min(q5, axis=0)                # 求各列最小值

    return max_uv[0], max_uv[1], min_uv[0], min_uv[1]


# 求相似度
def xsd(list3, list4):

    u1 = list3[0]
    u2 = list4[0]
    v1 = list3[1]
    v2 = list4[1]
    pai1 = 1 - u1 - v1
    pai2 = 1 - u2 - v2
    d1 = (pai1 + pai2) / 2
    d2 = abs(2 * (u1 - u2) - (v1 - v2)) / 3
    d3 = abs(2 * (v1 - v2) - (u1 - u2)) / 3
    s = 1 - d2 * (1 - d1) - d3 * d1

    return s


# 权重值计算
def qz(jz3):

    for c23 in range(0, len(jz3)):
        for c24 in range(0, 6):
            u = jz3[c23][c24][0]
            v = jz3[c23][c24][1]
            if u == 0:
                d1 = 0
            else:
                d1 = u * math.log(u)
            if v == 0:
                d2 = 0
            else:
                d2 = v * math.log(v)
            if u + v == 0:
                d3 = 0
            else:
                d3 = (u + v) * math.log(u + v)
            jz3[c23][c24] = d1 + d2 - d3 - (1 - u - v) * math.log(2)
    lqh = np.sum(np.array(jz3), axis=0)
    list5 = lqh.tolist()
    for c25 in range(0, len(list5)):
        list5[c25] = -1 / (6 * math.log(2)) * list5[c25]

    qh = np.sum(np.array(list5))
    for c26 in range(0, len(list5)):
        list5[c26] = (1 - list5[c26]) / (6 - qh)

    return list5


if __name__ == '__main__':

    a = [                                           # 威胁值列表
        [0.1, 0.2, 0.3, [0.4, 0.7], 0.1, 0.2],
        [0.2, 0.3, 0.5, [0.7, 1.0], 0.2, 0.3],
        [0.5, 0.7, 0.4, [0.0, 0.4], 0.7, 0.4]
    ]
    print('威胁值列表：')
    print(a)
    wb = op.load_workbook('data.xlsx')
    sh = wb["Sheet1"]
    for c19 in range(0, len(a)):
        for c20 in range(0, 6):
            sh.cell(c20 + 2, c19 + 2, str(a[c19][c20]))

    b = qjs_ifn(np.array(a))                        # 求区间数模糊数
    for c1 in range(0, len(a)):
        a[c1][3] = b[c1]
    c = ss_ifn(np.array(a), 0, 3)                   # 求前3列实数模糊数
    for c2 in range(0, len(a)):
        for c3 in range(0, 3):
            a[c2][c3] = c[c2][c3]
    d = ss_ifn(np.array(a), 4, 6)                   # 求后两列实数模糊数
    for c4 in range(0, len(a)):
        for c5 in range(4, 6):
            a[c4][c5] = d[c4][c5 - 4]
    print('直觉模糊数列表:')
    print(a)
    for c17 in range(0, len(a)):
        for c18 in range(0, 6):
            sh.cell(c18 + 15, c17 + 2, str(a[c17][c18]))
    a1 = np.array(a)
    a2 = a1.tolist()
    wk = qz(a2)                                     # 权重值列表
    print("权重值为：")
    print(wk)
    a1 = []                                         # 最优解列表
    a2 = []                                         # 最劣解列表
    for c5 in range(0, 6):                          # 求最优解和最劣解
        max_uij, max_vij, min_uij, min_vij = find_min_max(np.array(a), c5)
        a1.append([max_uij, min_vij])
        a2.append([min_uij, max_vij])
    print('最优解列表：')
    print(a1)
    print('最劣解列表：')
    print(a2)
    s1 = [[], [], []]                               # 根据坦克数量进行变化
    for c6 in range(0, len(a)):
        for c7 in range(0, 6):
            s1[c6].append(xsd(a1[c7], a[c6][c7]))
    print('最优解两两相似度列表：')
    print(s1)
    s2 = [[], [], []]                               # 根据坦克数量进行变化
    for c8 in range(0, len(a)):
        for c9 in range(0, 6):
            s2[c8].append(xsd(a2[c9], a[c8][c9]))
    print('最劣解两两相似度列表：')
    print(s2)

    s3 = []                                         # 最优解相似度列表
    s4 = []                                         # 最劣解相似度列表
    c12 = 0
    for c10 in range(0, len(a)):                    # 最优解相似度计算
        for c11 in range(0, 6):
            c12 = c12 + wk[c11] * s1[c10][c11]
        s3.append(c12)
        c12 = 0
    print('最优解相似度列表：')
    print(s3)
    for c13 in range(0, len(a)):                    # 最劣解相似度计算
        for c14 in range(0, 6):
            c12 = c12 + wk[c14] * s2[c13][c14]
        s4.append(c12)
        c12 = 0
    print('最劣解相似度列表：')
    print(s4)
    p = []                                          # 相对贴近度
    for c15 in range(0,len(a)):
        c16 = s3[c15] / (s3[c15] + s4[c15])
        p.append(c16)
    print('相对贴进度列表：')
    print(p)
    for c21 in range(0, len(a)):
        sh.cell(c21 + 28, 2, p[c21])

    array = np.array(p)
    px = np.argsort(-array)                         # 进行排序
    px = px.tolist()
    print('排序列表：')
    print(px)
    for c22 in range(0,len(a)):
        sh.cell(c22+28, 3, px[c22] + 1)
    print('排序结果：' + 'T' + str(px[0] + 1) + '>T' + str(px[1] + 1) + '>T' + str(px[2] + 1))
    wb.save("data.xlsx")
