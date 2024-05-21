import numpy as np
import openpyxl as op
import math


# 无量纲化处理
def wlg(jz):

    max_dimension1 = np.max(jz, axis=0)                   # 求各列最大值
    min_dimension1 = np.min(jz, axis=0)                   # 求各列最小值
    for c1 in range(0, jz.shape[0]):                      # 进行无量纲化处理
        for c2 in range(0, jz.shape[1]):
            if c2 == 1 | c2 == 2 | c2 == 3 | c2 == 4:
                if (max_dimension1[c2] - min_dimension1[c2]) != 0:
                    jz[c1, c2] = (jz[c1, c2] - min_dimension1[c2]) / (max_dimension1[c2] - min_dimension1[c2])
            else:
                if (max_dimension1[c2] - min_dimension1[c2]) != 0:
                    jz[c1, c2] = (max_dimension1[c2] - jz[c1, c2]) / (max_dimension1[c2] - min_dimension1[c2])
    list1 = jz.tolist()

    return list1


# 整体平移
# aef表示平移的数值
def py(jz1, aef):

    for c3 in range(0, len(jz1)):                         # 进行平移
        for c4 in range(0, 6):
            jz1[c3][c4] = jz1[c3][c4] + aef

    return jz1


# 特征比重计算
def tz_bz(jz2):

    lqh = np.sum(np.array(jz2), axis=0)                   # 求各列之和
    list2 = lqh.tolist()
    for c5 in range(0, len(jz2)):                         # 进行计算
        for c6 in range(0, 6):
            jz2[c5][c6] = jz2[c5][c6] / list2[c6]

    return jz2


# 熵值计算
def sz(jz3):

    for c7 in range(0, len(jz3)):
        for c8 in range(0, 6):
            jz3[c7][c8] = jz3[c7][c8] * math.log(jz3[c7][c8])
    lqh1 = np.sum(np.array(jz3), axis=0)                  # 求各列之和
    print(lqh1)
    list3 = lqh1.tolist()
    list4 = []
    for c9 in range(0, 6):                                # 进行计算
        list4.append(-1 / math.log(10) * list3[c9])

    return list4


# 权重计算
def qz(jz4):

    for c10 in range(0, 6):
        jz4[c10] = 1 - jz4[c10]
        jz7 = np.array(jz4)
        jz8 = jz7.tolist()
    qh = np.sum(np.array(jz4))                            # 求和

    for c11 in range(0, 6):                               # 进行计算
        jz4[c11] = jz4[c11] / qh

    return jz8, jz4


# 综合威胁度计算
def zh_wxd(jz5, jz6):

    list5 = []
    c14 = 0
    for c12 in range(0, len(jz6)):                       # 进行计算
        for c13 in range(0, 6):
            c14 = c14 + jz5[c13] * jz6[c12][c13]
        list5.append(c14)
        c14 = 0

    return list5


if __name__ == '__main__':

    a = [                                               # 威胁值列表
        [0.1, 0.2, 0.3, 0.5, 0.1, 0.2],
        [0.2, 0.3, 0.5, 0.7, 0.2, 0.3],
        [0.5, 0.7, 0.4, 0.1, 0.7, 0.4]
    ]
    print('威胁值列表：')
    print(a)
    wb = op.load_workbook('data1.xlsx')
    sh = wb["Sheet1"]
    for c15 in range(0, len(a)):
        for c16 in range(0, 6):
            sh.cell(c16 + 2, c15 + 2, a[c15][c16])
    b = wlg(np.array(a))                               # 无量纲化处理
    print('无量纲化列表：')
    print(b)
    for c17 in range(0, len(a)):
        for c18 in range(0, 6):
            sh.cell(c18 + 11, c17 + 2, b[c17][c18])
    c = py(b, 0.0001)                                  # 整体向右平移0.0001
    print('整体向右平移：')
    print(c)
    for c19 in range(0, len(a)):
        for c20 in range(0, 6):
            sh.cell(c20 + 20, c19 + 2, c[c19][c20])
    pij = tz_bz(c)                                     # 求特征比重
    print('特征比重列表：')
    print(pij)
    for c21 in range(0, len(a)):
        for c22 in range(0, 6):
            sh.cell(c22 + 29, c21 + 2, pij[c21][c22])

    pij1 = np.array(pij)
    pij2 = pij1.tolist()
    ej = sz(pij)                                        # 求熵值
    print('熵值列表:')
    print(ej)
    for c23 in range(0, 6):
        sh.cell(c23 + 38, 2, ej[c23])
    gj, wj = qz(ej)                                     # 求权重
    print('差异系数:')
    print(gj)
    for c24 in range(0, 6):
        sh.cell(c24 + 38, 3, gj[c24])
    print('权重值列表:')
    print(wj)
    for c25 in range(0, 6):
        sh.cell(c25 + 38, 4, wj[c25])
    zwd = zh_wxd(wj, pij2)                             # 求综合威胁度
    print('综合威胁度列表：')
    print(zwd)
    for c26 in range(0, len(a)):
        sh.cell(c26 + 47, 2, zwd[c26])
    array = np.array(zwd)
    px = np.argsort(-array)                            # 进行排序
    px = px.tolist()
    print('排序列表：')
    print(px)
    for c27 in range(0, len(a)):
        sh.cell(c27 + 47, 3, px[c27] + 1)
    print('排序结果：' + 'T' + str(px[0] + 1) + '>T' + str(px[1] + 1) + '>T' + str(px[2] + 1))
    wb.save("data1.xlsx")
